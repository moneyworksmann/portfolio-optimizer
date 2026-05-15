#!/usr/bin/env python3
"""
Portfolio Optimizer & Dashboard Generator
Fetches real-time prices, performs optimization, generates visualizations, and updates Excel
Run this script to refresh your portfolio analysis in real-time
"""

import yfinance as yf
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
from scipy.optimize import minimize
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from PIL import Image
import io
import warnings
warnings.filterwarnings('ignore')

# ============ CONFIGURATION ============
EXCEL_FILE = 'Investment_Portfolio_Tracker.xlsx'

SECTORS = {
    'IWM': 'Diversified', 'AAPL': 'Technology', 'AMZN': 'Consumer', 'JEPI': 'Fixed Income',
    'SLG': 'Real Estate', 'ORCL': 'Technology', 'MSFT': 'Technology', 'C': 'Financials',
    'JPM': 'Financials', 'MU': 'Technology', 'WBD': 'Consumer', 'LYG': 'Financials',
    'BAC': 'Financials', 'DAL': 'Industrials', 'NVDA': 'Technology', 'WMT': 'Consumer',
    'NTGR': 'Technology', 'T': 'Utilities'
}

HOLDINGS = {
    'IWM': 0.379, 'AAPL': 0.023, 'AMZN': 0.02, 'JEPI': 0.841,
    'SLG': 1, 'ORCL': 1, 'MSFT': 0.8, 'C': 1,
    'JPM': 1, 'MU': 0.55, 'WBD': 5, 'LYG': 10,
    'BAC': 1, 'DAL': 1, 'NVDA': 1, 'WMT': 1,
    'NTGR': 1, 'T': 1
}

PURCHASE_PRICES = {
    'IWM': 263.83, 'AAPL': 255.87, 'AMZN': 248.20, 'JEPI': 59.46,
    'SLG': 43.39, 'ORCL': 175.84, 'MSFT': 326.38, 'C': 95.95,
    'JPM': 304.68, 'MU': 153.54, 'WBD': 27.75, 'LYG': 3.70,
    'BAC': 45.88, 'DAL': 43.81, 'NVDA': 126.56, 'WMT': 66.99,
    'NTGR': 14.27, 'T': 16.49
}

# ============ FUNCTIONS ============

def fetch_prices(tickers):
    """Fetch real-time prices from Yahoo Finance"""
    print("📊 Fetching real-time prices from Yahoo Finance...")
    try:
        data = yf.download(tickers, period='1d', progress=False)
        if isinstance(data.columns, pd.MultiIndex):
            # Multiple tickers - use 'Close' from MultiIndex
            prices = data['Close'].iloc[-1].to_dict()
        else:
            # Single ticker - data is already a Series
            prices = {tickers[0]: data['Close'].iloc[-1]}
        print("✅ Prices fetched successfully!")
        return prices
    except Exception as e:
        print(f"⚠️  Error fetching prices: {e}")
        print("Using purchase prices as fallback...")
        return PURCHASE_PRICES.copy()

def calculate_portfolio(holdings, prices, purchase_prices):
    """Calculate portfolio metrics"""
    portfolio_values = {}
    portfolio_costs = {}
    gains = {}

    for ticker, shares in holdings.items():
        current_price = prices.get(ticker, purchase_prices[ticker])
        cost = shares * purchase_prices[ticker]
        value = shares * current_price
        portfolio_costs[ticker] = cost
        portfolio_values[ticker] = value
        gains[ticker] = value - cost

    total_value = sum(portfolio_values.values())
    total_cost = sum(portfolio_costs.values())
    total_gain = total_value - total_cost

    return {
        'values': portfolio_values, 'costs': portfolio_costs, 'gains': gains,
        'total_value': total_value, 'total_cost': total_cost, 'total_gain': total_gain
    }

def optimize_portfolio(tickers, prices, holdings, purchase_prices, cov_matrix, mean_returns):
    """Optimize portfolio using Modern Portfolio Theory"""
    print("\n📈 Running portfolio optimization...")

    def portfolio_stats(weights, mean_ret, cov_mat):
        ret = np.sum(mean_ret * weights) * 252
        risk = np.sqrt(np.dot(weights.T, np.dot(cov_mat * 252, weights)))
        sharpe = ret / risk if risk > 0 else 0
        return ret, risk, sharpe

    def neg_sharpe(weights, mean_ret, cov_mat):
        return -portfolio_stats(weights, mean_ret, cov_mat)[2]

    constraints = {'type': 'eq', 'fun': lambda w: np.sum(w) - 1}
    bounds = tuple((0, 0.2) for _ in tickers)
    initial_guess = np.array([1/len(tickers)] * len(tickers))

    result = minimize(neg_sharpe, initial_guess, args=(mean_returns, cov_matrix),
                     method='SLSQP', bounds=bounds, constraints=constraints)

    return result.x

def create_visualizations(portfolio_data, prices, current_weights, optimal_weights, tickers,
                         mean_returns, cov_matrix, hist_data):
    """Create all portfolio visualization charts"""
    print("\n🎨 Creating visualizations...")

    images = {}

    # Set style
    sns.set_style("whitegrid")
    plt.rcParams['figure.figsize'] = (12, 7)

    # 1. SECTOR CONCENTRATION PIE CHART
    fig, ax = plt.subplots(figsize=(10, 8))
    sector_values = {}
    for ticker, value in portfolio_data['values'].items():
        sector = SECTORS.get(ticker, 'Other')
        sector_values[sector] = sector_values.get(sector, 0) + value

    colors = plt.cm.Set3(np.linspace(0, 1, len(sector_values)))
    wedges, texts, autotexts = ax.pie(sector_values.values(), labels=sector_values.keys(),
                                        autopct='%1.1f%%', colors=colors, startangle=90)
    for autotext in autotexts:
        autotext.set_color('white')
        autotext.set_fontweight('bold')
        autotext.set_fontsize(10)
    ax.set_title('Portfolio Allocation by Sector', fontsize=14, fontweight='bold', pad=20)

    buf = io.BytesIO()
    plt.savefig(buf, format='png', dpi=150, bbox_inches='tight')
    buf.seek(0)
    images['sector'] = buf
    plt.close()

    # 2. CURRENT VS OPTIMAL ALLOCATION
    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(16, 8))

    top_current = sorted(zip(tickers, current_weights), key=lambda x: x[1], reverse=True)[:10]
    top_optimal = sorted(zip(tickers, optimal_weights), key=lambda x: x[1], reverse=True)[:10]

    ax1.barh([t[0] for t in top_current], [t[1]*100 for t in top_current], color='#FF6B6B')
    ax1.set_xlabel('Allocation %', fontsize=11, fontweight='bold')
    ax1.set_title('Current Allocation (Top 10)', fontsize=12, fontweight='bold')
    ax1.invert_yaxis()

    ax2.barh([t[0] for t in top_optimal], [t[1]*100 for t in top_optimal], color='#4ECDC4')
    ax2.set_xlabel('Allocation %', fontsize=11, fontweight='bold')
    ax2.set_title('Optimal Allocation (Top 10)', fontsize=12, fontweight='bold')
    ax2.invert_yaxis()

    plt.tight_layout()
    buf = io.BytesIO()
    plt.savefig(buf, format='png', dpi=150, bbox_inches='tight')
    buf.seek(0)
    images['allocation'] = buf
    plt.close()

    # 3. PORTFOLIO GROWTH CHART
    fig, ax = plt.subplots(figsize=(14, 7))

    # Calculate cumulative returns
    cumulative_returns = (hist_data / hist_data.iloc[0] - 1) * 100

    for ticker in tickers[:8]:  # Top 8 for clarity
        if ticker in cumulative_returns.columns:
            ax.plot(cumulative_returns.index, cumulative_returns[ticker], label=ticker, linewidth=2.5)

    ax.set_xlabel('Date', fontsize=11, fontweight='bold')
    ax.set_ylabel('Return (%)', fontsize=11, fontweight='bold')
    ax.set_title('Individual Stock Performance (1-Year)', fontsize=14, fontweight='bold')
    ax.legend(loc='best', fontsize=10)
    ax.grid(True, alpha=0.3)
    ax.axhline(y=0, color='black', linestyle='--', linewidth=1)

    plt.tight_layout()
    buf = io.BytesIO()
    plt.savefig(buf, format='png', dpi=150, bbox_inches='tight')
    buf.seek(0)
    images['growth'] = buf
    plt.close()

    # 4. SENSITIVITY ANALYSIS - What if different allocations?
    fig, ax = plt.subplots(figsize=(12, 7))

    allocation_scenarios = {
        'Current': current_weights,
        'Optimal': optimal_weights,
        'Equal Weight': np.array([1/len(tickers)] * len(tickers)),
        'Tech Heavy (30%)': np.array([0.30 if SECTORS.get(t) == 'Technology' else 0.70/(len(tickers)-5) for t in tickers]),
        'Conservative': np.array([0.10 if SECTORS.get(t) in ['Technology', 'Consumer'] else 0.90/(len(tickers)-7) for t in tickers])
    }

    scenario_returns = []
    for name, weights in allocation_scenarios.items():
        ret = np.sum(mean_returns * weights) * 252
        scenario_returns.append(ret * 100)

    colors_bar = ['#FF6B6B', '#4ECDC4', '#45B7D1', '#FFA07A', '#98D8C8']
    bars = ax.bar(allocation_scenarios.keys(), scenario_returns, color=colors_bar, alpha=0.8, edgecolor='black', linewidth=2)

    for bar in bars:
        height = bar.get_height()
        ax.text(bar.get_x() + bar.get_width()/2., height,
                f'{height:.1f}%', ha='center', va='bottom', fontsize=11, fontweight='bold')

    ax.set_ylabel('Expected Annual Return (%)', fontsize=11, fontweight='bold')
    ax.set_title('Sensitivity Analysis: Expected Returns by Allocation Strategy', fontsize=14, fontweight='bold')
    ax.grid(True, alpha=0.3, axis='y')

    plt.xticks(rotation=15, ha='right')
    plt.tight_layout()
    buf = io.BytesIO()
    plt.savefig(buf, format='png', dpi=150, bbox_inches='tight')
    buf.seek(0)
    images['sensitivity'] = buf
    plt.close()

    # 5. RISK-RETURN SCATTER
    fig, ax = plt.subplots(figsize=(12, 8))

    for ticker, weight in zip(tickers, current_weights):
        if weight > 0.01:
            ret = mean_returns.get(ticker, 0) * 252 * 100
            risk = np.sqrt(cov_matrix.loc[ticker, ticker] * 252) * 100
            size = weight * 1000
            ax.scatter(risk, ret, s=size, alpha=0.6, label=ticker)
            ax.annotate(ticker, (risk, ret), fontsize=9, ha='center')

    ax.set_xlabel('Risk (Volatility %)', fontsize=11, fontweight='bold')
    ax.set_ylabel('Return (Annual %)', fontsize=11, fontweight='bold')
    ax.set_title('Risk-Return Profile by Position (bubble size = allocation %)', fontsize=14, fontweight='bold')
    ax.grid(True, alpha=0.3)

    plt.tight_layout()
    buf = io.BytesIO()
    plt.savefig(buf, format='png', dpi=150, bbox_inches='tight')
    buf.seek(0)
    images['risk_return'] = buf
    plt.close()

    print("✅ Visualizations created!")
    return images

def update_excel(excel_file, portfolio_data, prices, current_weights, optimal_weights, tickers,
                mean_returns, cov_matrix, hist_data, images):
    """Update Excel file with current prices and visualizations"""
    print(f"\n📝 Updating {excel_file}...")

    wb = load_workbook(excel_file)

    # Update Holdings sheet with current prices
    holdings_sheet = wb['Holdings']
    for row_idx, ticker in enumerate(tickers, 2):
        current_price = prices.get(ticker, PURCHASE_PRICES[ticker])
        holdings_sheet[f'E{row_idx}'] = current_price

    # Create/update Optimization sheet with KPIs and charts
    if 'Optimization' in wb.sheetnames:
        opt_sheet = wb['Optimization']
        wb.remove(opt_sheet)

    opt_sheet = wb.create_sheet('Optimization', 3)

    # Title
    opt_sheet['A1'] = '📊 PORTFOLIO OPTIMIZATION DASHBOARD'
    opt_sheet['A1'].font = opt_sheet['A1'].font.copy()
    opt_sheet['A1'].font = opt_sheet['A1'].font.copy()

    # KPI Metrics
    opt_sheet['A3'] = 'KEY METRICS'
    opt_sheet['A4'] = f'Portfolio Value: ${portfolio_data["total_value"]:,.2f}'
    opt_sheet['A5'] = f'Total Gain/Loss: ${portfolio_data["total_gain"]:,.2f}'
    opt_sheet['A6'] = f'Return %: {(portfolio_data["total_gain"]/portfolio_data["total_cost"])*100:.2f}%'
    opt_sheet['A7'] = f'Updated: {pd.Timestamp.now().strftime("%Y-%m-%d %H:%M:%S")}'

    def portfolio_stats(weights, mean_ret, cov_mat):
        ret = np.sum(mean_ret * weights) * 252
        risk = np.sqrt(np.dot(weights.T, np.dot(cov_mat * 252, weights)))
        sharpe = ret / risk if risk > 0 else 0
        return ret, risk, sharpe

    current_ret, current_risk, current_sharpe = portfolio_stats(current_weights, mean_returns, cov_matrix)
    optimal_ret, optimal_risk, optimal_sharpe = portfolio_stats(optimal_weights, mean_returns, cov_matrix)

    opt_sheet['A9'] = 'PORTFOLIO METRICS'
    opt_sheet['A10'] = f'Current Annual Return: {current_ret*100:.2f}%'
    opt_sheet['A11'] = f'Current Risk (Volatility): {current_risk*100:.2f}%'
    opt_sheet['A12'] = f'Current Sharpe Ratio: {current_sharpe:.2f}'

    opt_sheet['A14'] = 'OPTIMAL ALLOCATION METRICS'
    opt_sheet['A15'] = f'Optimal Annual Return: {optimal_ret*100:.2f}%'
    opt_sheet['A16'] = f'Optimal Risk (Volatility): {optimal_risk*100:.2f}%'
    opt_sheet['A17'] = f'Optimal Sharpe Ratio: {optimal_sharpe:.2f}'
    opt_sheet['A18'] = f'Improvement in Sharpe: {((optimal_sharpe/current_sharpe - 1) * 100):.1f}%'

    # Insert images
    row_offset = 20
    if 'sector' in images:
        img = XLImage(images['sector'])
        img.width = 350
        img.height = 300
        opt_sheet.add_image(img, f'A{row_offset}')
        row_offset += 20

    if 'allocation' in images:
        img = XLImage(images['allocation'])
        img.width = 450
        img.height = 300
        opt_sheet.add_image(img, f'A{row_offset}')
        row_offset += 20

    if 'growth' in images:
        img = XLImage(images['growth'])
        img.width = 450
        img.height = 300
        opt_sheet.add_image(img, f'A{row_offset}')
        row_offset += 20

    if 'sensitivity' in images:
        img = XLImage(images['sensitivity'])
        img.width = 450
        img.height = 300
        opt_sheet.add_image(img, f'A{row_offset}')
        row_offset += 20

    if 'risk_return' in images:
        img = XLImage(images['risk_return'])
        img.width = 450
        img.height = 300
        opt_sheet.add_image(img, f'A{row_offset}')

    wb.save(excel_file)
    print(f"✅ Excel file updated with real-time data and visualizations!")

def main():
    print("\n" + "="*60)
    print("🚀 PORTFOLIO OPTIMIZER & DASHBOARD GENERATOR")
    print("="*60)

    tickers = list(HOLDINGS.keys())

    # Fetch prices
    prices = fetch_prices(tickers)

    # Calculate portfolio
    portfolio_data = calculate_portfolio(HOLDINGS, prices, PURCHASE_PRICES)
    print(f"\n💰 Portfolio Value: ${portfolio_data['total_value']:,.2f}")
    print(f"📈 Total Gain: ${portfolio_data['total_gain']:,.2f} ({(portfolio_data['total_gain']/portfolio_data['total_cost'])*100:.2f}%)")

    # Get historical data for optimization
    print("\n📊 Fetching historical data for optimization...")
    try:
        data = yf.download(tickers, period='1y', progress=False)
        if isinstance(data.columns, pd.MultiIndex):
            hist_data = data['Close']  # Use 'Close' instead of 'Adj Close'
        else:
            hist_data = data['Close'] if 'Close' in data.columns else data
        returns = hist_data.pct_change().dropna()
        mean_returns = returns.mean()
        cov_matrix = returns.cov()
    except Exception as e:
        print(f"⚠️  Using mock historical data for optimization... ({e})")
        mean_returns = pd.Series({t: 0.15 for t in tickers})
        cov_matrix = pd.DataFrame(np.random.rand(len(tickers), len(tickers)),
                                 index=tickers, columns=tickers)
        hist_data = pd.DataFrame(np.random.randn(252, len(tickers)), columns=tickers).cumsum()

    # Calculate weights
    current_weights = np.array([portfolio_data['values'][t] / portfolio_data['total_value'] for t in tickers])
    optimal_weights = optimize_portfolio(tickers, prices, HOLDINGS, PURCHASE_PRICES, cov_matrix, mean_returns)

    # Create visualizations
    images = create_visualizations(portfolio_data, prices, current_weights, optimal_weights,
                                  tickers, mean_returns, cov_matrix, hist_data)

    # Update Excel
    update_excel(EXCEL_FILE, portfolio_data, prices, current_weights, optimal_weights,
                tickers, mean_returns, cov_matrix, hist_data, images)

    print("\n" + "="*60)
    print("✅ PORTFOLIO UPDATE COMPLETE!")
    print("="*60)
    print(f"\nYour portfolio has been updated with:")
    print("  ✓ Real-time prices from Yahoo Finance")
    print("  ✓ Portfolio optimization analysis")
    print("  ✓ Sector concentration charts")
    print("  ✓ Growth performance visualization")
    print("  ✓ Sensitivity analysis")
    print("  ✓ Risk-return profile")
    print(f"\n📂 Updated file: {EXCEL_FILE}")
    print("\nRun this script weekly to keep your analysis current!")
    print("="*60 + "\n")

if __name__ == '__main__':
    main()
